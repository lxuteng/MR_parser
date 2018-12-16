import openpyxl
import os
import sys
import csv
import sqlite3
import pandas
import math
import time
from concurrent.futures import ThreadPoolExecutor as TPE
import traceback
import multiprocessing

################################################################################
# Module multiprocessing is organized differently in Python 3.4+
try:
    # Python 3.4+
    if sys.platform.startswith('win'):
        import multiprocessing.popen_spawn_win32 as forking
    else:
        import multiprocessing.popen_fork as forking
except ImportError:
    import multiprocessing.forking as forking

if sys.platform.startswith('win'):
    # First define a modified version of Popen.
    class _Popen(forking.Popen):
        def __init__(self, *args, **kw):
            if hasattr(sys, 'frozen'):
                # We have to set original _MEIPASS2 value from sys._MEIPASS
                # to get --onefile mode working.
                os.putenv('_MEIPASS2', sys._MEIPASS)
            try:
                super(_Popen, self).__init__(*args, **kw)
            finally:
                if hasattr(sys, 'frozen'):
                    # On some platforms (e.g. AIX) 'os.unsetenv()' is not
                    # available. In those cases we cannot delete the variable
                    # but only set it to the empty string. The bootloader
                    # can handle this case.
                    if hasattr(os, 'unsetenv'):
                        os.unsetenv('_MEIPASS2')
                    else:
                        os.putenv('_MEIPASS2', '')


    # Second override 'Popen' class with our modified version.
    forking.Popen = _Popen


################################################################################

class Main:
    def __init__(self):
        self.main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
        self.dt_point_list = []

    def get_dt_info(self):
        path_base_data = os.path.join(self.main_path, 'DT采样点_饶平.xlsx')
        f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
        for temp_sheet_name in f_base_data_wb.sheetnames:
            if temp_sheet_name == 'DT采样点':
                temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
                for temp_row in temp_f_base_data_wb_sheet.iter_rows(min_row=2):
                    temp_value = [str(j.value) for j in temp_row]
                    self.dt_point_list.append(temp_value)

    def get_mdt_info(self):
        path_base_data = os.path.join(self.main_path, 'raoping.csv')
        # self.conn = sqlite3.connect('db.db',check_same_thread=False)
        self.conn = sqlite3.connect(':memory:', check_same_thread=False)
        df = pandas.read_csv(path_base_data, encoding='gbk')
        df.to_sql('mdt', self.conn, if_exists='append', index=False)

    def executer(self, point_list):
        try:
            long, lat = point_list
            cu = self.conn.cursor()
            minlng, maxlng, minlat, maxlat = self.get_area(float(long), float(lat), 5)
            temp_sql_scr = self.sql_scr.replace('&1', minlng).replace('&2', maxlng).replace('&3', minlat).replace('&4',
                                                                                                                  maxlat)
            cu.execute(temp_sql_scr)
            # self.temp_list['value']['_'.join(point_list)] = cu.fetchall()
            self.temp_list['value'].append(point_list + list(cu.fetchall()[0]))
        except:
            pass
            # traceback.print_exc()

    def executer_head(self, point_list):
        long, lat = point_list
        cu = self.conn.cursor()
        minlng, maxlng, minlat, maxlat = self.get_area(float(long), float(lat), 5)
        temp_sql_scr = self.sql_scr.replace('&1', minlng).replace('&2', maxlng).replace('&3', minlat).replace('&4',
                                                                                                              maxlat)
        cu.execute(temp_sql_scr)
        self.temp_list['head'] = [i[0] for i in cu.description]

    def progress(self):
        self.get_mdt_info()

        self.temp_list = {
            'head': [],
            'value': []
        }
        f = open(os.path.join(self.main_path, 'sql.sql'), encoding='utf-8-sig')
        self.sql_scr = f.read()
        with TPE(1) as executor:
            executor.map(self.executer, self.dt_point_list)
        self.executer_head(self.dt_point_list[1])
        self.writer()

    def progress_process_obj(self,value_list):
        self.get_mdt_info()
        self.temp_list = {
            'head': [],
            'value': []
        }
        f = open(os.path.join(self.main_path, 'sql.sql'), encoding='utf-8-sig')
        self.sql_scr = f.read()
        for temp_value in value_list:
            self.executer(temp_value)
        self.executer_head(value_list[1])
        self.writer()


    def progress_process(self):
        core_num = 3
        dt_point_list_num = int(len(self.dt_point_list) / core_num)
        temp_dt_point_list = [self.dt_point_list[0:dt_point_list_num]]
        for i in range(core_num):
            if i+2 < core_num:
                temp_dt_point_list.append(self.dt_point_list[dt_point_list_num*(i+1):dt_point_list_num * (i+2)])
            elif i+2 == core_num:
                temp_dt_point_list.append(self.dt_point_list[dt_point_list_num*(i+1):])

        process_pool = multiprocessing.Pool(core_num)
        for temp_list in temp_dt_point_list:
            process_pool.apply_async(self.progress_process_obj, args=(temp_list,))
        process_pool.close()
        process_pool.join()

    def get_area(self, longitude, latitude, dis):
        """
        确定查询经纬度范围
        :param latitude:中心纬度
        :param longitude:中心经度
        :param dis:半径(m)
        :return:(minlat, maxlat, minlng, maxlng)
        """
        r = 6371.137
        dlng = 2 * math.asin(math.sin(dis / 1000 / (2 * r)) / math.cos(latitude * math.pi / 180))
        dlng = dlng * 180 / math.pi

        dlat = dis / 1000 / r
        dlat = dlat * 180 / math.pi

        minlng = longitude - dlng
        maxlng = longitude + dlng
        minlat = latitude - dlat
        maxlat = latitude + dlat

        return str(minlng), str(maxlng), str(minlat), str(maxlat)

    def writer(self):
        if not os.path.exists(os.path.join(self.main_path, 'converge_list.csv')):
            with open(os.path.join(self.main_path, 'converge_list.csv'), 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['DT_Longitude', 'DT_Latitude'] + self.temp_list['head'])
                writer.writerows(self.temp_list['value'])
        else:
            n = 1
            while n:
                try:
                    with open(os.path.join(self.main_path, 'converge_list.csv'), 'a', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerows(self.temp_list['value'])
                        n = 0
                except:
                    time.sleep(1)


if __name__ == '__main__':
    multiprocessing.freeze_support()

    print(time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))

    main = Main()
    main.get_dt_info()
    star_time = time.time()

    # main.progress()
    main.progress_process()

    print('>>> 历时：', time.strftime('%Y/%m/%d %H:%M:%S', time.gmtime(time.time() - star_time)))
    print(time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))
